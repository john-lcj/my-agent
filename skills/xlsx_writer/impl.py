"""xlsx_writer skill:把 rows 或 CSV 文本写成 Excel(.xlsx)。"""
from __future__ import annotations

import csv
import io
import os

from core.types import CapabilityResult
from governance.workspace import resolve_path

RISK = "WRITE"

SCHEMA = {
    "type": "object",
    "properties": {
        "path": {"type": "string", "description": "Output .xlsx path"},
        "rows": {"type": "array", "description": "Two-dimensional rows array; use either rows or csv_text"},
        "csv_text": {"type": "string", "description": "CSV text; use either csv_text or rows"},
        "sheet": {"type": "string", "description": "Worksheet name; defaults to Sheet1"},
    },
    "required": ["path"],
}


async def run(args: dict, ctx) -> CapabilityResult:
    path, error = resolve_path(str(args.get("path", "")).strip())
    if error:
        return CapabilityResult(ok=False, error=error)
    if not path:
        return CapabilityResult(ok=False, error="缺少参数 path")
    if not path.endswith(".xlsx"):
        path += ".xlsx"

    rows = args.get("rows")
    if not rows:
        csv_text = str(args.get("csv_text", "")).strip()
        if not csv_text:
            return CapabilityResult(ok=False, error="需提供 rows 或 csv_text")
        rows = list(csv.reader(io.StringIO(csv_text)))
    if not isinstance(rows, list) or not rows:
        return CapabilityResult(ok=False, error="rows 为空或格式不对")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as e:
        return CapabilityResult(ok=False, error=f"需要 openpyxl:{e}")

    wb = Workbook()
    ws = wb.active
    ws.title = str(args.get("sheet", "") or "Sheet1")[:31]
    for r, row in enumerate(rows, start=1):
        cells = row if isinstance(row, (list, tuple)) else [row]
        for c, val in enumerate(cells, start=1):
            ws.cell(row=r, column=c, value=val)
    for cell in ws[1]:  # 表头加粗
        cell.font = Font(bold=True)

    try:
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        wb.save(path)
    except Exception as e:
        return CapabilityResult(ok=False, error=f"写入失败:{e}")
    ncol = max((len(r) if isinstance(r, (list, tuple)) else 1) for r in rows)
    return CapabilityResult(ok=True, output=f"已生成 Excel:{path}({len(rows)} 行 × {ncol} 列)")
